#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ZhangCalibrationApp - 张氏相机标定及结果验证桌面应用（Fluent Design）

使用 PyQt-Fluent-Widgets 实现 Windows 11 Fluent 风格，按钮带悬停/按压动画，
布局 7:3，全局间距 15px、内边距 20px，卡片圆角 12px + 柔和阴影。
"""

import os
import sys

def _setup_qt_plugin_path():
    try:
        import PyQt5
        base = os.path.dirname(PyQt5.__file__)
        for sub in ("Qt", "Qt5"):
            plugin_dir = os.path.join(base, sub, "plugins", "platforms")
            if os.path.isdir(plugin_dir):
                os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = plugin_dir
                break
    except Exception:
        pass
_setup_qt_plugin_path()

import glob
import numpy as np
import cv2
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QHBoxLayout,
    QVBoxLayout,
    QGridLayout,
    QStackedWidget,
    QListWidget,
    QListWidgetItem,
    QLabel,
    QPushButton,
    QProgressBar,
    QFileDialog,
    QMessageBox,
    QGroupBox,
    QFormLayout,
    QFrame,
    QScrollArea,
    QButtonGroup,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QHeaderView,
    QSizePolicy,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSize, QPropertyAnimation, QEasingCurve, QRect, pyqtProperty, QTimer
from PyQt5.QtGui import QPixmap, QImage, QIcon, QFont, QColor

# Fluent 控件（若不可用则回退到标准控件）
try:
    from qfluentwidgets import (
        setTheme,
        Theme,
        PushButton,
        PrimaryPushButton,
        SpinBox,
        DoubleSpinBox,
        ComboBox,
    )
    _FLUENT_AVAILABLE = True
except ImportError:
    from PyQt5.QtWidgets import QSpinBox as SpinBox, QDoubleSpinBox as DoubleSpinBox, QComboBox as ComboBox
    PushButton = PrimaryPushButton = QPushButton
    _FLUENT_AVAILABLE = False

# 打包成 .app 后，__file__ 不可用：SCRIPT_DIR 用于文件对话框默认路径，path_for_import 用于 sys.path
if getattr(sys, "frozen", False):
    SCRIPT_DIR = os.path.expanduser("~")
    _path_for_import = os.path.dirname(sys.executable)
    _icon_dir = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    _path_for_import = SCRIPT_DIR
    _icon_dir = SCRIPT_DIR
if _path_for_import not in sys.path:
    sys.path.insert(0, _path_for_import)

def _app_icon_path():
    """应用图标路径（Mac 用 .icns，Windows 用 .ico；开发时用项目目录，打包后用 bundle 内路径）。"""
    if sys.platform == "win32":
        p = os.path.join(_icon_dir, "icon.ico")
    else:
        p = os.path.join(_icon_dir, "icon.icns")
    return p if os.path.isfile(p) else None

import camera_calibration as calib
import image_undistort as undistort

# ------------------------------ 全局常量 ------------------------------
LAYOUT_SPACING = 15
CONTAINER_MARGIN = 20
CARD_RADIUS = 12
BG_COLOR = "rgb(32, 32, 32)"
CARD_COLOR = "rgb(45, 45, 45)"
FONT_FAMILY = "Microsoft YaHei UI"
FONT_SIZE = 11

# ------------------------------ Fluent 主题 + 自定义样式 ------------------------------
def _apply_fluent_style(app):
    if _FLUENT_AVAILABLE:
        setTheme(Theme.DARK)
    # 全局字体与背景
    font = QFont(FONT_FAMILY, FONT_SIZE)
    app.setFont(font)
    APP_STYLESHEET = f"""
    QMainWindow {{ background-color: {BG_COLOR}; }}
    QWidget {{ background: transparent; color: #e0e0e0; font-family: "{FONT_FAMILY}"; font-size: {FONT_SIZE}pt; }}
    QGroupBox {{
        font-weight: 600; border: none; border-radius: {CARD_RADIUS}px;
        margin-top: 12px; padding: 16px;
        background-color: {CARD_COLOR};
    }}
    QGroupBox::title {{ subcontrol-origin: margin; left: 16px; padding: 0 8px; color: #b4a7f5; }}
    QLabel {{ color: #e0e0e0; }}
    QListWidget {{
        background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px;
        padding: 8px; outline: none;
    }}
    QListWidget::item {{ padding: 10px 12px; border-radius: 8px; margin: 2px 0; }}
    QListWidget::item:selected {{ background-color: #5e5cc9; color: white; }}
    QListWidget::item:hover:!selected {{ background-color: rgb(55,55,55); }}
    QScrollArea {{ border: none; background: transparent; }}
    QPushButton:checked {{ background-color: #5e5cc9; color: white; }}
    QPushButton#tabBtn {{
        min-width: 140px; padding: 10px 20px; border: none; border-radius: 8px 8px 0 0;
        background-color: rgb(50,50,50); color: #c0c0c0;
    }}
    QPushButton#tabBtn:hover {{ background-color: rgb(60,60,60); color: #e0e0e0; }}
    QPushButton#tabBtn:checked {{ background-color: #5e5cc9; color: white; font-weight: 600; }}
    QPushButton#primaryBtn {{
        background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #7c6cf5, stop:1 #5e5cc9);
        color: white; border: none; border-radius: 8px;
    }}
    QPushButton#primaryBtn:hover {{ background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #8d7ef7, stop:1 #7c6cf5); }}
    QPushButton#primaryBtn:pressed {{ background: #4a49a8; }}
    """
    app.setStyleSheet(APP_STYLESHEET)


def _add_card_shadow(widget):
    """为深色卡片添加柔和阴影，替代硬边框。"""
    from PyQt5.QtWidgets import QGraphicsDropShadowEffect
    from PyQt5.QtGui import QColor
    shadow = QGraphicsDropShadowEffect(widget)
    shadow.setBlurRadius(24)
    shadow.setXOffset(0)
    shadow.setYOffset(4)
    shadow.setColor(QColor(0, 0, 0, 80))
    widget.setGraphicsEffect(shadow)


# ------------------------------ 带动画的按钮包装器 ------------------------------
class AnimatedButtonWrapper(QWidget):
    """
    包装任意 QPushButton（或 Fluent PushButton），提供：
    - 悬停：200ms 内缩放至 1.05
    - 按压：整体缩小至 0.98
    """
    def __init__(self, button, parent=None):
        super().__init__(parent)
        self._btn = button
        self._base_w = max(80, button.sizeHint().width())
        self._base_h = max(36, button.sizeHint().height())
        self._scale_val = 1.0
        self._target_scale = 1.0
        self._anim = QPropertyAnimation(self, b"scale")
        self._anim.setDuration(200)
        self._anim.setEasingCurve(QEasingCurve.OutCubic)
        self._anim.valueChanged.connect(self._apply_scale)
        self.setFixedSize(int(self._base_w * 1.08), int(self._base_h * 1.08))
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignCenter)
        layout.addWidget(button)
        button.setFixedSize(self._base_w, self._base_h)
        button.installEventFilter(self)

    def get_scale(self):
        return self._scale_val
    def set_scale(self, v):
        self._scale_val = v
        self._apply_scale()
    scale = pyqtProperty(float, get_scale, set_scale)

    def _apply_scale(self):
        s = self._scale_val
        w = int(self._base_w * s)
        h = int(self._base_h * s)
        self._btn.setFixedSize(w, h)
        x = (self.width() - w) // 2
        y = (self.height() - h) // 2
        self._btn.move(x, y)

    def _animate_to(self, target):
        if abs(self._target_scale - target) < 0.001:
            return
        self._target_scale = target
        self._anim.stop()
        self._anim.setStartValue(self._scale_val)
        self._anim.setEndValue(target)
        self._anim.start()

    def eventFilter(self, obj, event):
        if obj is self._btn:
            from PyQt5.QtCore import QEvent
            t = event.type()
            if t == QEvent.Enter:
                self._animate_to(1.05)
            elif t == QEvent.Leave:
                self._animate_to(1.0)
            elif t == QEvent.MouseButtonPress:
                self._animate_to(0.98)
            elif t == QEvent.MouseButtonRelease:
                self._animate_to(1.05 if self._btn.underMouse() else 1.0)
        return super().eventFilter(obj, event)

    def button(self):
        return self._btn


# ------------------------------ 异步标定线程 ------------------------------
class CalibrationWorker(QThread):
    """后台执行 camera_calibration.run_calibration。"""
    finished = pyqtSignal(object)
    error = pyqtSignal(str)
    log_line = pyqtSignal(str)
    progress = pyqtSignal(int, int)  # (current, total)

    def __init__(self, image_dir, pattern_size, square_size_mm, save_output=False):
        super().__init__()
        self.image_dir = image_dir
        self.pattern_size = pattern_size
        self.square_size_mm = square_size_mm
        self.save_output = save_output

    def run(self):
        try:
            def on_log(s):
                self.log_line.emit(s)
            def on_progress(current, total):
                self.progress.emit(current, total)
            result = calib.run_calibration(
                self.image_dir,
                pattern_size=self.pattern_size,
                square_size_mm=self.square_size_mm,
                output_file=None,
                output_dir=None,
                log_callback=on_log,
                progress_callback=on_progress,
                save_output=self.save_output,
            )
            self.finished.emit(result)
        except Exception as e:
            self.log_line.emit(f"错误: {e}")
            self.error.emit(str(e))
            self.finished.emit(None)


# ------------------------------ 异步矫正+验证线程 ------------------------------
class UndistortWorker(QThread):
    """后台执行 image_undistort.run_undistort_and_verify。"""
    finished = pyqtSignal(object)
    progress = pyqtSignal(int, int)  # (current, total) - 用于动态进度条
    status_message = pyqtSignal(str)  # 当前步骤说明，用于界面显示

    def __init__(self, image_path, calib_path, pattern_hint, square_size_mm, tolerance_mm=0.1):
        super().__init__()
        self.image_path = image_path
        self.calib_path = calib_path
        self.pattern_hint = pattern_hint
        self.square_size_mm = square_size_mm
        self.tolerance_mm = tolerance_mm

    def run(self):
        try:
            self.progress.emit(0, 0)
            def on_status(msg):
                self.status_message.emit(msg)
            out = undistort.run_undistort_and_verify(
                self.image_path,
                self.calib_path,
                pattern_hint=self.pattern_hint,
                square_size_mm=self.square_size_mm,
                tolerance_mm=self.tolerance_mm,
                output_dir=None,
                status_callback=on_status,
            )
            self.finished.emit(out)
        except Exception as e:
            self.status_message.emit("出错: " + str(e))
            self.finished.emit({"error": str(e)})


def _ndarray_to_pixmap(bgr, max_size=(800, 600)):
    if bgr is None or bgr.size == 0:
        return None
    h, w = bgr.shape[:2]
    if len(bgr.shape) == 2:
        img = cv2.cvtColor(bgr, cv2.COLOR_GRAY2RGB)
    else:
        img = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
    scale = min(max_size[0] / w, max_size[1] / h, 1.0)
    if scale < 1.0:
        nw, nh = int(w * scale), int(h * scale)
        img = cv2.resize(img, (nw, nh), interpolation=cv2.INTER_AREA)
    h, w = img.shape[:2]
    bytes_per_line = w * 3
    qimg = QImage(img.data, w, h, bytes_per_line, QImage.Format_RGB888)
    return QPixmap.fromImage(qimg.copy())


def _show_info(parent, title, text):
    """提示弹框：确保文字与背景对比清晰（深色文字+浅色背景）。"""
    mb = QMessageBox(parent)
    mb.setWindowTitle(title)
    mb.setText(text)
    mb.setIcon(QMessageBox.Information)
    mb.setStyleSheet(
        "QMessageBox { background-color: #f5f5f0; } "
        "QLabel { color: #1e1e1e; } "
        "QPushButton { color: #1e1e1e; background-color: #d0d0c8; }"
    )
    mb.exec_()


def _show_warning(parent, title, text):
    mb = QMessageBox(parent)
    mb.setWindowTitle(title)
    mb.setText(text)
    mb.setIcon(QMessageBox.Warning)
    mb.setStyleSheet(
        "QMessageBox { background-color: #f5f5f0; } "
        "QLabel { color: #1e1e1e; } "
        "QPushButton { color: #1e1e1e; background-color: #d0d0c8; }"
    )
    mb.exec_()


def _show_critical(parent, title, text):
    mb = QMessageBox(parent)
    mb.setWindowTitle(title)
    mb.setText(text)
    mb.setIcon(QMessageBox.Critical)
    mb.setStyleSheet(
        "QMessageBox { background-color: #f5f5f0; } "
        "QLabel { color: #1e1e1e; } "
        "QPushButton { color: #1e1e1e; background-color: #d0d0c8; }"
    )
    mb.exec_()


# ------------------------------ 可鼠标滚轮缩放 + 拖拽平移的图像显示控件 ------------------------------
class ZoomableImageWidget(QScrollArea):
    """支持鼠标滚轮缩放、鼠标拖拽平移，便于查看细节与整体。支持进度条显示。"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._pixmap = None
        self._scale = 1.0
        self._min_scale = 0.2
        self._max_scale = 5.0
        self._drag_start = None  # QPoint，拖拽起始位置
        self.setWidgetResizable(True)  # 改为True，让容器可以调整大小
        self.setAlignment(Qt.AlignCenter)
        
        # 创建容器widget
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(0)
        
        self._label = QLabel()
        self._label.setAlignment(Qt.AlignCenter)
        self._label.setScaledContents(False)  # 不使用自动缩放，手动控制
        self._label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # 允许扩展
        self._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; min-height: 200px;")
        container_layout.addWidget(self._label, stretch=1)
        
        # 进度条（默认隐藏）
        self._progress_bar = QProgressBar()
        self._progress_bar.setVisible(False)
        self._progress_bar.setStyleSheet(f"""
            QProgressBar {{
                border: none;
                border-radius: 4px;
                background-color: rgb(50, 50, 50);
                height: 6px;
                text-align: center;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5e5cc9, stop:0.5 #7c6cf5, stop:1 #5e5cc9);
                border-radius: 4px;
            }}
        """)
        container_layout.addWidget(self._progress_bar)
        
        self.setWidget(container)

    def set_pixmap(self, pixmap):
        self._pixmap = pixmap
        self._scale = 1.0  # 重置缩放比例
        self._drag_start = None
        if pixmap is None or pixmap.isNull():
            self.hide_progress()  # 隐藏进度条
            self._update_display()
        else:
            # 延迟更新显示，确保widget已经有大小
            QTimer.singleShot(10, self._update_display)

    def set_progress(self, current, total):
        """设置进度条：如果total>0则显示真实进度，否则显示动态流动进度"""
        if total > 0:
            self._progress_bar.setRange(0, total)
            self._progress_bar.setValue(current)
            self._progress_bar.setVisible(True)
        else:
            # 动态流动进度条
            self._progress_bar.setRange(0, 0)  # 设置为不确定模式
            self._progress_bar.setVisible(True)
    
    def hide_progress(self):
        """隐藏进度条"""
        self._progress_bar.setVisible(False)
    
    def _update_display(self):
        if self._pixmap is None or self._pixmap.isNull():
            self._label.clear()
            self._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; min-height: 200px;")
            self._label.setMinimumSize(0, 200)  # 重置最小大小
            return
        
        # 获取可用空间
        viewport_size = self.viewport().size()
        available_width = max(viewport_size.width(), 100)  # 至少100像素
        available_height = max(viewport_size.height() - (self._progress_bar.height() if self._progress_bar.isVisible() else 0), 100)
        
        # 计算适合可用空间的缩放比例
        pixmap_width = self._pixmap.width()
        pixmap_height = self._pixmap.height()
        
        if pixmap_width <= 0 or pixmap_height <= 0:
            return
        
        # 计算初始缩放比例（如果scale=1.0，则适应可用空间）
        if self._scale == 1.0:
            scale_w = available_width / pixmap_width
            scale_h = available_height / pixmap_height
            fit_scale = min(scale_w, scale_h)  # 适应可用空间，可以放大或缩小
            nw = int(pixmap_width * fit_scale)
            nh = int(pixmap_height * fit_scale)
        else:
            # 使用用户设置的缩放比例
            nw = int(pixmap_width * self._scale)
            nh = int(pixmap_height * self._scale)
        
        # 确保至少有一些大小
        nw = max(nw, 100)
        nh = max(nh, 100)
        
        scaled = self._pixmap.scaled(nw, nh, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        
        self._label.setPixmap(scaled)
        # 设置label的大小，确保图像能够完全显示
        self._label.setMinimumSize(scaled.size())
        self._label.resize(scaled.size())
        self._label.setMaximumSize(16777215, 16777215)  # QWIDGETSIZE_MAX
        self._label.setStyleSheet("")  # 清除样式，让图像正常显示
        self.hide_progress()  # 图像显示时隐藏进度条

    def wheelEvent(self, event):
        delta = event.angleDelta().y()
        if delta > 0:
            self._scale = min(self._max_scale, self._scale * 1.15)
        else:
            self._scale = max(self._min_scale, self._scale / 1.15)
        self._update_display()
        event.accept()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = event.pos()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_start is not None and event.buttons() & Qt.LeftButton:
            delta = event.pos() - self._drag_start
            self._drag_start = event.pos()
            h = self.horizontalScrollBar()
            v = self.verticalScrollBar()
            h.setValue(h.value() - delta.x())
            v.setValue(v.value() - delta.y())
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_start = None
        super().mouseReleaseEvent(event)
    
    def resizeEvent(self, event):
        """窗口大小变化时，如果scale=1.0则重新适应空间"""
        super().resizeEvent(event)
        if self._pixmap is not None and not self._pixmap.isNull() and self._scale == 1.0:
            QTimer.singleShot(10, self._update_display)


# ------------------------------ 标定与结果（同一界面，7:3 网格 + Fluent） ------------------------------
class CalibrationAndResultsPage(QWidget):
    """
    标定数据准备 + 结果报表：QGridLayout 左侧图像区 7、右侧参数区 3，
    所有按钮带悬停/按压动画，卡片圆角 12px + 阴影。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._image_paths = []
        self._image_dir = ""
        self._pattern_size = (0, 0)
        self._square_size_mm = 0.0
        self._calibration_done = False  # 仅点击「开始标定」完成后，预览才显示角点
        self._worker = None
        self._last_calib_result = None  # 标定完成后保存在内存中的结果（未自动保存文件时用此生成 npz）
        self._setup_ui()

    def _setup_ui(self):
        grid = QGridLayout(self)
        grid.setSpacing(LAYOUT_SPACING)
        grid.setContentsMargins(CONTAINER_MARGIN, CONTAINER_MARGIN, CONTAINER_MARGIN, CONTAINER_MARGIN)

        # 左列：仅目录卡片，占满整列高度（列表在卡片内拉伸）
        left_col = QVBoxLayout()
        left_col.setContentsMargins(0, 0, 0, 0)
        left_card = QFrame()
        left_card.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px;")
        left_card.setFixedWidth(240)
        _add_card_shadow(left_card)
        left_layout = QVBoxLayout(left_card)
        left_layout.setContentsMargins(CONTAINER_MARGIN // 2, CONTAINER_MARGIN // 2, CONTAINER_MARGIN // 2, CONTAINER_MARGIN // 2)
        left_layout.setSpacing(LAYOUT_SPACING)
        title = QLabel("标定图片")
        title.setStyleSheet(f"font-size: 12pt; font-weight: 600; color: #b4a7f5;")
        left_layout.addWidget(title)
        hint = QLabel("选择包含棋盘格的图片目录")
        hint.setStyleSheet("color: #a0a0a0; font-size: 10pt;")
        hint.setWordWrap(True)
        left_layout.addWidget(hint)
        btn_open = PushButton("选择标定图片目录") if _FLUENT_AVAILABLE else QPushButton("选择标定图片目录")
        btn_open.setCursor(Qt.PointingHandCursor)
        btn_open.clicked.connect(self._on_open_dir)
        left_layout.addWidget(AnimatedButtonWrapper(btn_open))
        self._label_calib_dir = QLabel("未选择目录")
        self._label_calib_dir.setWordWrap(True)
        self._label_calib_dir.setStyleSheet(f"color: #a0a0a0; font-size: 10pt; padding: 8px; border-radius: 8px; background-color: rgb(38,38,38);")
        left_layout.addWidget(self._label_calib_dir)
        self._list_widget = QListWidget()
        self._list_widget.setIconSize(QSize(80, 60))
        self._list_widget.setSpacing(6)
        self._list_widget.currentRowChanged.connect(self._on_selection_changed)
        left_layout.addWidget(self._list_widget, 1)  # 列表拉伸填满卡片剩余高度
        left_col.addWidget(left_card, 1)  # 卡片拉伸填满左列整列
        grid.addLayout(left_col, 0, 0)
        grid.setColumnStretch(0, 2)

        # 中列：仅图像预览 + 标题 + 检测状态，预览居中不挤到下方
        center_col = QVBoxLayout()
        center_col.setSpacing(LAYOUT_SPACING)
        center_col.addWidget(QLabel("图像预览（滚轮缩放；点击「开始标定」后显示角点）"))
        self._preview_area = ZoomableImageWidget()
        self._preview_area.setMinimumSize(400, 300)
        self._preview_label = self._preview_area._label
        self._preview_label.setText("还未上传图像")
        self._preview_label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #a0a0a0; font-size: 12pt;")
        center_col.addWidget(self._preview_area, 1)  # 预览占满中间剩余空间，内容在 ZoomableImageWidget 内居中
        self._label_detect_status = QLabel("检测结果: --")
        self._label_detect_status.setStyleSheet(f"padding: 12px; border-radius: 8px; background-color: {CARD_COLOR}; color: #b4a7f5; font-weight: 600;")
        center_col.addWidget(self._label_detect_status)
        grid.addLayout(center_col, 0, 1)
        grid.setColumnStretch(1, 7)

        # 右侧列：参数 + 开始标定 + 结果（比例 3）
        right_col = QVBoxLayout()
        right_col.setSpacing(LAYOUT_SPACING)
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setStyleSheet("QScrollArea { background: transparent; border: none; }")
        right_w = QWidget()
        right_layout = QVBoxLayout(right_w)
        right_layout.setSpacing(LAYOUT_SPACING)

        g_params = QGroupBox("棋盘格参数")
        g_params.setStyleSheet(f"QGroupBox {{ background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; }}")
        _add_card_shadow(g_params)
        form = QFormLayout()
        self._spin_cols = SpinBox()
        self._spin_cols.setRange(0, 20)
        self._spin_cols.setValue(0)
        self._spin_cols.valueChanged.connect(self._on_params_changed)
        self._spin_rows = SpinBox()
        self._spin_rows.setRange(0, 20)
        self._spin_rows.setValue(0)
        self._spin_rows.valueChanged.connect(self._on_params_changed)
        self._spin_square_mm = DoubleSpinBox()
        self._spin_square_mm.setRange(0.0, 100.0)
        self._spin_square_mm.setValue(0.0)
        self._spin_square_mm.setSuffix(" mm")
        self._spin_square_mm.valueChanged.connect(self._on_params_changed)
        def _tit(t):
            l = QLabel(t)
            l.setStyleSheet("color: #b4a7f5;")
            return l
        form.addRow(_tit("内角点列数:"), self._spin_cols)
        form.addRow(_tit("内角点行数:"), self._spin_rows)
        form.addRow(_tit("格子边长 (mm):"), self._spin_square_mm)
        g_params.setLayout(form)
        right_layout.addWidget(g_params)

        self._btn_calibrate = PrimaryPushButton("开始标定") if _FLUENT_AVAILABLE else QPushButton("开始标定")
        if not _FLUENT_AVAILABLE:
            self._btn_calibrate.setObjectName("primaryBtn")
        self._btn_calibrate.setCursor(Qt.PointingHandCursor)
        self._btn_calibrate.setFixedHeight(44)
        self._btn_calibrate.clicked.connect(self._on_start_calibration)
        right_layout.addWidget(AnimatedButtonWrapper(self._btn_calibrate))
        self._progress = QProgressBar()
        self._progress.setVisible(False)
        right_layout.addWidget(self._progress)

        g_log = QGroupBox("运行日志")
        g_log.setStyleSheet(f"QGroupBox {{ background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; }}")
        _add_card_shadow(g_log)
        log_layout = QVBoxLayout()
        self._log_edit = QTextEdit()
        self._log_edit.setReadOnly(True)
        self._log_edit.setPlaceholderText("点击「开始标定」后此处显示角点检测与标定过程…")
        self._log_edit.setStyleSheet(f"font-family: Consolas; font-size: 10pt; color: #c0b8d8; background-color: rgb(38,38,38); border: none; border-radius: 8px;")
        self._log_edit.setMaximumHeight(160)
        log_layout.addWidget(self._log_edit)
        g_log.setLayout(log_layout)
        right_layout.addWidget(g_log)

        g_result = QGroupBox("结果报表")
        g_result.setStyleSheet(f"QGroupBox {{ background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; }}")
        _add_card_shadow(g_result)
        result_layout = QVBoxLayout()
        self._table_result = QTableWidget()
        self._table_result.setColumnCount(2)
        self._table_result.setHorizontalHeaderLabels(["参数", "值"])
        self._table_result.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self._table_result.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self._table_result.setRowCount(0)
        self._table_result.setStyleSheet(f"QTableWidget {{ background-color: transparent; color: #c0b8d8; gridline-color: rgb(60,60,60); }}")
        self._table_result.verticalHeader().setVisible(False)
        self._table_result.setMaximumHeight(280)
        result_layout.addWidget(self._table_result)
        self._btn_download_npz = PushButton("下载 npz 到本地")
        self._btn_download_npz.setCursor(Qt.PointingHandCursor)
        self._btn_download_npz.clicked.connect(self._on_download_npz)
        self._btn_download_npz.setEnabled(False)
        result_layout.addWidget(AnimatedButtonWrapper(self._btn_download_npz))
        g_result.setLayout(result_layout)
        right_layout.addWidget(g_result)
        right_layout.addStretch(1)
        right_scroll.setWidget(right_w)
        right_col.addWidget(right_scroll)
        grid.addLayout(right_col, 0, 2)
        grid.setColumnStretch(2, 3)

    def _on_open_dir(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择标定图片所在目录", SCRIPT_DIR)
        if not dir_path:
            return
        self._image_dir = dir_path
        self._label_calib_dir.setText(dir_path)
        exts = ("*.jpg", "*.JPG", "*.png", "*.PNG", "*.HEIC", "*.heic")
        paths = []
        for ext in exts:
            paths.extend(glob.glob(os.path.join(dir_path, ext)))
        self._image_paths = sorted(set(paths))
        self._list_widget.clear()
        for p in self._image_paths:
            item = QListWidgetItem(os.path.basename(p))
            item.setData(Qt.UserRole, p)
            img = calib.read_image(p)
            if img is not None and img.size > 0:
                pix = _ndarray_to_pixmap(img, max_size=(120, 90))
                if pix is not None:
                    item.setIcon(QIcon(pix))
            self._list_widget.addItem(item)
        if not self._image_paths:
            _show_info(self, "提示", "该目录下未找到支持的图片。")
        else:
            self._list_widget.setCurrentRow(0)

    def _on_params_changed(self):
        self._pattern_size = (self._spin_cols.value(), self._spin_rows.value())
        self._square_size_mm = self._spin_square_mm.value()
        self._refresh_preview()

    def _on_selection_changed(self, row):
        if 0 <= row < len(self._image_paths):
            self._refresh_preview()

    def _refresh_preview(self):
        row = self._list_widget.currentRow()
        if row < 0 or row >= len(self._image_paths):
            self._preview_area.set_pixmap(None)
            self._preview_area.hide_progress()  # 隐藏进度条
            self._preview_label.setText("还未上传图像")
            self._preview_label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #a0a0a0; font-size: 12pt;")
            self._label_detect_status.setText("检测结果: --")
            return
        path = self._image_paths[row]
        img = calib.read_image(path)
        if img is None:
            self._preview_area.set_pixmap(None)
            self._preview_label.setText("无法读取图像")
            self._preview_label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #ff6b6b; font-size: 12pt;")
            self._label_detect_status.setText("检测结果: 无法读取图像")
            return
        # 仅点击「开始标定」完成后才在预览中显示角点，否则只显示原图
        if not self._calibration_done or self._pattern_size[0] < 3 or self._pattern_size[1] < 3:
            pix = _ndarray_to_pixmap(img, max_size=(1200, 900))
            if pix is not None:
                self._preview_area.set_pixmap(pix)
            self._label_detect_status.setText("检测结果: 请先点击「开始标定」后显示角点" if not self._calibration_done else "检测结果: 请设置有效棋盘格参数")
            return
        ret, corners, _ = calib.find_corners(img, self._pattern_size)
        if ret and corners is not None:
            vis = img.copy()
            cv2.drawChessboardCorners(vis, self._pattern_size, corners, True)
            pix = _ndarray_to_pixmap(vis, max_size=(1200, 900))
        else:
            pix = _ndarray_to_pixmap(img, max_size=(1200, 900))
        if pix is not None:
            self._preview_area.set_pixmap(pix)
            self._label_detect_status.setText("检测结果: 已检测到角点" if ret else "检测结果: 未检测到角点")
        else:
            self._label_detect_status.setText("检测结果: 未检测到角点")

    def _on_start_calibration(self):
        if not self._image_paths:
            _show_warning(self, "提示", "请先选择标定图片目录。")
            return
        cols, rows = self._spin_cols.value(), self._spin_rows.value()
        square_size_mm = self._spin_square_mm.value()
        if cols < 3 or rows < 3:
            _show_warning(self, "提示", "内角点列数、行数均需 ≥ 3，请设置有效的棋盘格参数。")
            return
        if square_size_mm <= 0:
            _show_warning(self, "提示", "格子边长 (mm) 需大于 0。")
            return
        image_dir = os.path.dirname(self._image_paths[0])
        pattern_size = (cols, rows)
        self._log_edit.clear()
        self._btn_calibrate.setEnabled(False)
        self._progress.setVisible(True)
        self._progress.setRange(0, 0)
        # 设置预览区域显示"正在处理图片"状态
        self._preview_area.set_pixmap(None)
        self._preview_label.setText("正在处理图片...")
        self._preview_label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #b4a7f5; font-size: 12pt;")
        self._preview_area.set_progress(0, 0)  # 显示动态流动进度条
        self._worker = CalibrationWorker(image_dir, pattern_size, square_size_mm, save_output=False)
        self._worker.finished.connect(self._on_calibration_finished)
        self._worker.error.connect(self._on_calibration_error)
        self._worker.log_line.connect(self._on_calibration_log)
        self._worker.progress.connect(self._on_calibration_progress)
        self._worker.start()

    def _on_calibration_log(self, line):
        self._log_edit.append(line)
    
    def _on_calibration_progress(self, current, total):
        """更新标定进度"""
        self._preview_area.set_progress(current, total)

    def _on_calibration_finished(self, result):
        self._progress.setVisible(False)
        self._preview_area.hide_progress()  # 隐藏进度条
        self._btn_calibrate.setEnabled(True)
        if result is None:
            self._preview_label.setText("标定失败，请重试")
            self._preview_label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #ff6b6b; font-size: 12pt;")
            return
        self._calibration_done = True
        self._last_calib_result = result  # 保存到内存，供「下载 npz 到本地」使用
        self._pattern_size = (self._spin_cols.value(), self._spin_rows.value())
        self._square_size_mm = self._spin_square_mm.value()
        self._fill_result(result)
        self._btn_download_npz.setEnabled(True)
        self._refresh_preview()
        _show_info(self, "标定完成", "标定已完成，结果已显示在右侧。可点击「下载 npz 到本地」保存。")

    def _on_calibration_error(self, msg):
        self._progress.setVisible(False)
        self._btn_calibrate.setEnabled(True)
        self._preview_label.setText("标定失败，请重试")
        self._preview_label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #ff6b6b; font-size: 12pt;")
        _show_critical(self, "标定错误", msg)

    def _fill_result(self, result):
        mtx = result.get("camera_matrix")
        dist = result.get("dist_coeffs")
        err = result.get("mean_error")
        rows = []
        if mtx is not None:
            rows.extend([
                ("fx", f"{mtx[0, 0]:.6f}"),
                ("fy", f"{mtx[1, 1]:.6f}"),
                ("cx", f"{mtx[0, 2]:.6f}"),
                ("cy", f"{mtx[1, 2]:.6f}"),
            ])
        if dist is not None:
            d = dist.ravel()
            names = ["k1", "k2", "p1", "p2", "k3"]
            for i in range(min(5, len(d))):
                rows.append((names[i], f"{d[i]:.6f}"))
        if err is not None:
            rows.append(("平均重投影误差 (像素)", f"{err:.4f}"))
        self._table_result.setRowCount(len(rows))
        for r, (name, val) in enumerate(rows):
            self._table_result.setItem(r, 0, QTableWidgetItem(name))
            self._table_result.setItem(r, 1, QTableWidgetItem(val))

    def _on_download_npz(self):
        # 标定完成后结果在内存中（_last_calib_result），不再依赖磁盘上的 npz 文件
        if not self._last_calib_result:
            _show_warning(self, "提示", "暂无标定结果，请先完成标定。")
            return
        default_name = "calibration_result.npz"
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 npz 到本地",
            default_name,
            "NumPy (*.npz)"
        )
        if not path:
            return
        try:
            r = self._last_calib_result
            np.savez(
                path,
                ret=r.get("ret"),
                camera_matrix=r.get("camera_matrix"),
                dist_coeffs=r.get("dist_coeffs"),
                rvecs=r.get("rvecs"),
                tvecs=r.get("tvecs"),
                image_size=r.get("image_size"),
                mean_error=r.get("mean_error"),
            )
            _show_info(self, "提示", f"已保存到：{path}")
        except Exception as e:
            _show_critical(self, "错误", str(e))


# ------------------------------ 畸变校正验证（严格按 image_undistort） ------------------------------
class UndistortionValidationPage(QWidget):
    """
    严格按 image_undistort：标定系数矫正图像；验证时在矫正前/矫正后图像上按横纵 1/5 与 4/5
    确定同一裁剪区域，在该区域内检测格点、在相邻点线段上标注像素当量，并统计
    最小/最大像素当量、差值、差/最小%。完整呈现 run_undistort_and_verify 的分析结果，并提供下载矫正图。
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._calib_path = None
        self._image_paths = []
        self._current_index = 0
        self._last_result = None
        self._worker = None
        self._all_results = []  # 存储所有图片的验证结果，用于生成Excel报告
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(LAYOUT_SPACING)
        layout.setContentsMargins(CONTAINER_MARGIN, CONTAINER_MARGIN, CONTAINER_MARGIN, CONTAINER_MARGIN)

        # 控制栏 + 路径显示
        ctrl = QHBoxLayout()
        ctrl.setSpacing(LAYOUT_SPACING)
        btn_calib = PushButton("加载标定 (npz)") if _FLUENT_AVAILABLE else QPushButton("加载标定 (npz)")
        btn_calib.setCursor(Qt.PointingHandCursor)
        btn_calib.clicked.connect(self._on_load_calib)
        ctrl.addWidget(AnimatedButtonWrapper(btn_calib))
        self._label_calib_path = QLabel("未加载标定")
        self._label_calib_path.setStyleSheet(f"color: #a0a0a0; font-size: 10pt; padding: 8px; border-radius: 8px; background-color: rgb(38,38,38);")
        self._label_calib_path.setWordWrap(True)
        ctrl.addWidget(self._label_calib_path, stretch=1)
        layout.addLayout(ctrl)

        ctrl2 = QHBoxLayout()
        ctrl2.setSpacing(LAYOUT_SPACING)
        btn_dir = PushButton("选择验证图片目录") if _FLUENT_AVAILABLE else QPushButton("选择验证图片目录")
        btn_dir.setCursor(Qt.PointingHandCursor)
        btn_dir.clicked.connect(self._on_open_images)
        ctrl2.addWidget(AnimatedButtonWrapper(btn_dir))
        self._label_image_dir = QLabel("未选择目录")
        self._label_image_dir.setStyleSheet(f"color: #a0a0a0; font-size: 10pt; padding: 8px; border-radius: 8px; background-color: rgb(38,38,38);")
        self._label_image_dir.setWordWrap(True)
        ctrl2.addWidget(self._label_image_dir, stretch=1)
        layout.addLayout(ctrl2)

        # 当前图片 与 显示 并列一行：左侧当前图片，右侧显示
        row_current_display = QHBoxLayout()
        row_current_display.setSpacing(LAYOUT_SPACING)
        lbl_cur = QLabel("当前图片:")
        lbl_cur.setStyleSheet("color: #b4a7f5;")
        row_current_display.addWidget(lbl_cur)
        self._combo_images = ComboBox()
        self._combo_images.setMinimumWidth(200)
        self._combo_images.currentIndexChanged.connect(self._on_image_changed)
        row_current_display.addWidget(self._combo_images)
        row_current_display.addSpacing(24)
        lbl_display = QLabel("显示:")
        lbl_display.setStyleSheet("color: #b4a7f5;")
        row_current_display.addWidget(lbl_display)
        self._combo_display = ComboBox()
        self._combo_display.addItem("整图（原图 / 矫正图）", "full")
        self._combo_display.addItem("裁剪验证图（矫正前 / 矫正后）", "crop")
        self._combo_display.currentIndexChanged.connect(self._on_display_mode_changed)
        self._combo_display.setMinimumWidth(220)
        row_current_display.addWidget(self._combo_display)
        layout.addLayout(row_current_display)

        # 主内容：左侧矫正前/矫正后左右并排（可缩放）+ 右侧格点验证竖排
        content = QHBoxLayout()
        content.setSpacing(LAYOUT_SPACING)
        left_panel = QVBoxLayout()
        left_panel.setSpacing(LAYOUT_SPACING)
        images_row = QHBoxLayout()
        images_row.setSpacing(LAYOUT_SPACING)
        left_img_col = QVBoxLayout()
        lbl_orig = QLabel("矫正前")
        lbl_orig.setStyleSheet("color: #b4a7f5;")
        left_img_col.addWidget(lbl_orig)
        self._zoom_orig = ZoomableImageWidget()
        self._zoom_orig.setMinimumSize(320, 280)
        self._zoom_orig._label.setText("还未上传图像")
        self._zoom_orig._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #a0a0a0; font-size: 12pt;")
        left_img_col.addWidget(self._zoom_orig, stretch=1)
        images_row.addLayout(left_img_col, stretch=1)
        right_img_col = QVBoxLayout()
        lbl_undist = QLabel("矫正后")
        lbl_undist.setStyleSheet("color: #b4a7f5;")
        right_img_col.addWidget(lbl_undist)
        self._zoom_undist = ZoomableImageWidget()
        self._zoom_undist.setMinimumSize(320, 280)
        self._zoom_undist._label.setText("还未上传图像")
        self._zoom_undist._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #a0a0a0; font-size: 12pt;")
        right_img_col.addWidget(self._zoom_undist, stretch=1)
        images_row.addLayout(right_img_col, stretch=1)
        left_panel.addLayout(images_row, stretch=1)
        content.addLayout(left_panel, 3)

        right_panel = QVBoxLayout()
        g_verify = QGroupBox("格点验证结果（横纵 1/5 与 4/5 裁剪区域）")
        g_verify.setStyleSheet(f"QGroupBox {{ background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; }}")
        _add_card_shadow(g_verify)
        verify_layout = QVBoxLayout()
        # 当前进行状态（矫正验证过程中实时更新）
        self._verify_status = QLabel()
        self._verify_status.setWordWrap(True)
        self._verify_status.setStyleSheet("color: #b4a7f5; font-weight: 600; padding: 6px 0; min-height: 22px;")
        self._verify_status.setText("")
        self._verify_status.setVisible(False)
        verify_layout.addWidget(self._verify_status)
        self._verify_text = QLabel()
        self._verify_text.setWordWrap(True)
        self._verify_text.setStyleSheet(f"font-family: Consolas; padding: 12px; background-color: transparent; color: #e0e0e0;")
        self._verify_text.setText("请加载标定并选择图片，将自动进行矫正与格点验证。")
        verify_layout.addWidget(self._verify_text)
        self._btn_download = PrimaryPushButton("下载矫正后图片到本地") if _FLUENT_AVAILABLE else QPushButton("下载矫正后图片到本地")
        if not _FLUENT_AVAILABLE:
            self._btn_download.setObjectName("primaryBtn")
        self._btn_download.setCursor(Qt.PointingHandCursor)
        self._btn_download.clicked.connect(self._on_download)
        self._btn_download.setEnabled(False)
        verify_layout.addWidget(AnimatedButtonWrapper(self._btn_download))
        g_verify.setLayout(verify_layout)
        right_panel.addWidget(g_verify)
        content.addLayout(right_panel, 1)
        layout.addLayout(content)

        param_row = QHBoxLayout()
        param_row.setSpacing(LAYOUT_SPACING)
        lbl_square = QLabel("格子边长 (mm):")
        lbl_square.setStyleSheet("color: #b4a7f5;")
        param_row.addWidget(lbl_square)
        self._spin_square = DoubleSpinBox()
        self._spin_square.setRange(0.1, 100.0)
        self._spin_square.setValue(5.0)
        self._spin_square.setSuffix(" mm")
        self._spin_square.valueChanged.connect(self._run_undistort_if_ready)
        param_row.addWidget(self._spin_square)
        param_row.addSpacing(LAYOUT_SPACING)
        lbl_tolerance = QLabel("容差 (mm):")
        lbl_tolerance.setStyleSheet("color: #b4a7f5;")
        param_row.addWidget(lbl_tolerance)
        self._spin_tolerance = DoubleSpinBox()
        self._spin_tolerance.setRange(0.01, 10.0)
        self._spin_tolerance.setValue(0.1)
        self._spin_tolerance.setSuffix(" mm")
        param_row.addWidget(self._spin_tolerance)
        param_row.addStretch(1)
        layout.addLayout(param_row)

        # 生成Excel报告按钮
        btn_row = QHBoxLayout()
        btn_row.setSpacing(LAYOUT_SPACING)
        self._btn_generate_report = PrimaryPushButton("生成矫正报告 (Excel)") if _FLUENT_AVAILABLE else QPushButton("生成矫正报告 (Excel)")
        if not _FLUENT_AVAILABLE:
            self._btn_generate_report.setObjectName("primaryBtn")
        self._btn_generate_report.setCursor(Qt.PointingHandCursor)
        self._btn_generate_report.clicked.connect(self._on_generate_report)
        self._btn_generate_report.setEnabled(False)
        btn_row.addWidget(AnimatedButtonWrapper(self._btn_generate_report))
        btn_row.addStretch(1)
        layout.addLayout(btn_row)

    def _on_load_calib(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择标定文件",
            os.path.join(SCRIPT_DIR, "train_result", "txt_result"),
            "NumPy 标定 (*.npz);;所有 (*.*)"
        )
        if not path:
            return
        self._calib_path = path
        self._label_calib_path.setText(path)
        _show_info(self, "提示", "标定文件已加载。请选择验证图片目录并选择图片。")
        self._run_undistort_if_ready()

    def _on_open_images(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择验证图片目录", SCRIPT_DIR)
        if not dir_path:
            return
        self._label_image_dir.setText(dir_path)
        exts = ("*.jpg", "*.JPG", "*.png", "*.PNG", "*.HEIC", "*.heic")
        paths = []
        for ext in exts:
            paths.extend(glob.glob(os.path.join(dir_path, ext)))
        self._image_paths = sorted(set(paths))
        self._combo_images.clear()
        for p in self._image_paths:
            self._combo_images.addItem(os.path.basename(p), p)
        if self._image_paths:
            self._combo_images.setCurrentIndex(0)
        # 清空之前的结果
        self._all_results = []
        self._btn_generate_report.setEnabled(False)

    def _on_display_mode_changed(self):
        self._update_display_images()

    def _update_display_images(self):
        """根据「整图/裁剪验证图」选择刷新左右两图（可缩放）。裁剪验证图必须使用带 crop_verify 标记的图。"""
        out = self._last_result
        if not out or out.get("error"):
            return
        # 用 currentIndex 判断裁剪模式，保证“裁剪验证图”一定用带标记的 crop_orig_annot / crop_undist_annot
        use_crop = (self._combo_display.currentIndex() == 1)
        if use_crop:
            img_orig = out.get("crop_orig_annot")
            img_undist = out.get("crop_undist_annot")
        else:
            path = self._image_paths[self._current_index] if self._current_index < len(self._image_paths) else None
            if path:
                img_orig = undistort.read_image(path)
            else:
                img_orig = None
            img_undist = out.get("undistorted")
        # 左右两侧都要更新，避免裁剪模式下仍残留整图
        if img_orig is not None:
            pix = _ndarray_to_pixmap(img_orig, max_size=(1200, 900))
            if pix is not None:
                self._zoom_orig.set_pixmap(pix)
                self._zoom_orig._label.setText("")
                self._zoom_orig.hide_progress()  # 隐藏进度条
        else:
            self._zoom_orig.set_pixmap(None)
            self._zoom_orig.hide_progress()  # 隐藏进度条
            self._zoom_orig._label.setText("还未上传图像" if not use_crop else "无裁剪验证图（矫正前）")
            self._zoom_orig._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #a0a0a0; font-size: 12pt;")
        if img_undist is not None:
            pix_u = _ndarray_to_pixmap(img_undist, max_size=(1200, 900))
            if pix_u is not None:
                self._zoom_undist.set_pixmap(pix_u)
                self._zoom_undist._label.setText("")
                self._zoom_undist.hide_progress()  # 隐藏进度条
        else:
            self._zoom_undist.set_pixmap(None)
            self._zoom_undist.hide_progress()  # 隐藏进度条
            self._zoom_undist._label.setText("还未上传图像" if not use_crop else "无裁剪验证图（矫正后）")
            self._zoom_undist._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #a0a0a0; font-size: 12pt;")

    def _on_image_changed(self, index):
        if 0 <= index < len(self._image_paths):
            self._current_index = index
            self._run_undistort_if_ready()

    def _run_undistort_if_ready(self):
        if not self._calib_path or not self._image_paths or self._current_index >= len(self._image_paths):
            return
        if self._worker is not None and self._worker.isRunning():
            return
        path = self._image_paths[self._current_index]
        self._verify_status.setVisible(True)
        self._verify_status.setText("正在矫正与验证…")
        self._verify_text.setText("")
        self._zoom_orig.set_pixmap(None)
        self._zoom_orig._label.setText("正在处理图片...")
        self._zoom_orig._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #b4a7f5; font-size: 12pt;")
        self._zoom_orig.set_progress(0, 0)  # 显示动态流动进度条
        self._zoom_undist.set_pixmap(None)
        self._zoom_undist._label.setText("正在处理图片...")
        self._zoom_undist._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #b4a7f5; font-size: 12pt;")
        self._zoom_undist.set_progress(0, 0)  # 显示动态流动进度条
        self._btn_download.setEnabled(False)
        self._worker = UndistortWorker(
            path, self._calib_path, None, self._spin_square.value(),
            tolerance_mm=self._spin_tolerance.value()
        )
        self._worker.finished.connect(self._on_undistort_finished)
        self._worker.progress.connect(self._on_undistort_progress)
        self._worker.status_message.connect(self._on_undistort_status)
        self._worker.start()

    def _on_undistort_progress(self, current, total):
        """更新矫正进度（动态进度条）"""
        self._zoom_orig.set_progress(current, total)
        self._zoom_undist.set_progress(current, total)

    def _on_undistort_status(self, msg):
        """更新界面上的当前进行状态"""
        self._verify_status.setText(msg)

    def _on_undistort_finished(self, out):
        self._zoom_orig.hide_progress()
        self._zoom_undist.hide_progress()
        self._last_result = out
        err = out.get("error")
        if err:
            self._verify_status.setVisible(False)
            self._verify_status.setText("")
            self._zoom_orig.set_pixmap(None)
            self._zoom_orig._label.setText("矫正失败")
            self._zoom_orig._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #ff6b6b; font-size: 12pt;")
            self._zoom_undist.set_pixmap(None)
            self._zoom_undist._label.setText(f"错误: {err}")
            self._zoom_undist._label.setStyleSheet(f"background-color: {CARD_COLOR}; border: none; border-radius: {CARD_RADIUS}px; color: #ff6b6b; font-size: 12pt;")
            self._verify_text.setText(f"矫正失败：{err}")
            return
        # 保存结果到_all_results用于生成Excel报告
        image_name = os.path.splitext(os.path.basename(self._image_paths[self._current_index]))[0] if self._current_index < len(self._image_paths) else "unknown"
        result_data = {
            "image_name": image_name,
            "image_path": self._image_paths[self._current_index] if self._current_index < len(self._image_paths) else None,
            "result_orig": out.get("result_orig"),
            "result_undist": out.get("result_undist"),
            "crop_orig_annot": out.get("crop_orig_annot"),
            "crop_undist_annot": out.get("crop_undist_annot"),
            "undistorted": out.get("undistorted"),
            "square_size_mm": self._spin_square.value(),
        }
        # 更新或添加当前图片的结果
        found = False
        for i, r in enumerate(self._all_results):
            if r["image_name"] == image_name:
                self._all_results[i] = result_data
                found = True
                break
        if not found:
            self._all_results.append(result_data)
        self._btn_generate_report.setEnabled(len(self._all_results) > 0)
        self._verify_status.setVisible(False)
        self._verify_status.setText("")
        self._verify_text.setText(out.get("summary_text", "无验证结果文本。"))
        self._btn_download.setEnabled(True)
        self._update_display_images()

    def _on_download(self):
        if not self._last_result or self._last_result.get("error"):
            _show_warning(self, "提示", "暂无矫正结果可下载。")
            return
        dir_path = QFileDialog.getExistingDirectory(self, "选择保存目录", SCRIPT_DIR)
        if not dir_path:
            return
        base = self._last_result.get("base_name", "undistorted")
        undist = self._last_result.get("undistorted")
        crop_orig = self._last_result.get("crop_orig_annot")
        crop_undist = self._last_result.get("crop_undist_annot")
        try:
            if undist is not None:
                cv2.imwrite(os.path.join(dir_path, f"{base}_undistorted.jpg"), undist)
            if crop_orig is not None:
                cv2.imwrite(os.path.join(dir_path, f"{base}_crop_original_verify.jpg"), crop_orig)
            if crop_undist is not None:
                cv2.imwrite(os.path.join(dir_path, f"{base}_crop_undistorted_verify.jpg"), crop_undist)
            _show_info(self, "提示", f"已保存到：{dir_path}")
        except Exception as e:
            _show_critical(self, "错误", str(e))

    def _on_generate_report(self):
        if not self._all_results:
            _show_warning(self, "提示", "暂无验证结果可生成报告。")
            return
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel报告",
            os.path.join(SCRIPT_DIR, "矫正验证报告.xlsx"),
            "Excel文件 (*.xlsx)"
        )
        if not file_path:
            return
        try:
            self._generate_excel_report(file_path)
            _show_info(self, "提示", f"Excel报告已生成：{file_path}")
        except Exception as e:
            _show_critical(self, "错误", f"生成Excel报告失败：{str(e)}")

    def _generate_excel_report(self, file_path):
        """生成Excel矫正验证报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage
        except ImportError:
            raise ImportError("需要安装 openpyxl 和 Pillow 库：pip install openpyxl Pillow")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "矫正验证报告"
        
        # 设置行高（为图片预留空间）
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, len(self._all_results) + 2):
            ws.row_dimensions[row_idx].height = 100
        
        # 设置列标题
        headers = ["工位", "矫正前图", "矫正后图", "检测项", "标准值 (mm)", 
                   "矫正前检出值 (mm)", "矫正后检出值 (mm)", "差值 (mm)", 
                   "容差 (mm)", "是否通过"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        
        tolerance = self._spin_tolerance.value()
        
        # 填充数据
        for row_idx, result in enumerate(self._all_results, 2):
            image_name = result["image_name"]
            square_size = result["square_size_mm"]
            result_orig = result["result_orig"]
            result_undist = result["result_undist"]
            
            # 工位（图像名称）
            ws.cell(row=row_idx, column=1, value=image_name)
            
            # 矫正前图、矫正后图（保存为临时文件并插入到Excel）
            temp_dir = os.path.join(os.path.dirname(file_path), "temp_images")
            os.makedirs(temp_dir, exist_ok=True)
            
            crop_orig = result.get("crop_orig_annot")
            crop_undist = result.get("crop_undist_annot")
            
            # 插入矫正前图
            if crop_orig is not None:
                orig_path = os.path.join(temp_dir, f"{image_name}_before.jpg")
                cv2.imwrite(orig_path, crop_orig)
                try:
                    img = XLImage(orig_path)
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f'B{row_idx}')
                except Exception:
                    ws.cell(row=row_idx, column=2, value=f"[图片: {image_name}_before.jpg]")
            else:
                ws.cell(row=row_idx, column=2, value="无")
            
            # 插入矫正后图
            if crop_undist is not None:
                undist_path = os.path.join(temp_dir, f"{image_name}_after.jpg")
                cv2.imwrite(undist_path, crop_undist)
                try:
                    img = XLImage(undist_path)
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f'C{row_idx}')
                except Exception:
                    ws.cell(row=row_idx, column=3, value=f"[图片: {image_name}_after.jpg]")
            else:
                ws.cell(row=row_idx, column=3, value="无")
            
            # 检测项
            ws.cell(row=row_idx, column=4, value="广角畸变")
            
            # 标准值
            ws.cell(row=row_idx, column=5, value=square_size)
            
            # 矫正前检出值（最大格子长度）
            if result_orig is not None:
                max_before = result_orig.get("max_length_mm", 0.0)
                ws.cell(row=row_idx, column=6, value=round(max_before, 3))
            else:
                ws.cell(row=row_idx, column=6, value="--")
            
            # 矫正后检出值（最大格子长度）
            if result_undist is not None:
                max_after = result_undist.get("max_length_mm", 0.0)
                ws.cell(row=row_idx, column=7, value=round(max_after, 3))
            else:
                ws.cell(row=row_idx, column=7, value="--")
            
            # 差值（矫正后检出值 - 标准值）
            if result_undist is not None:
                diff = abs(max_after - square_size)
                ws.cell(row=row_idx, column=8, value=round(diff, 3))
            else:
                ws.cell(row=row_idx, column=8, value="--")
            
            # 容差
            ws.cell(row=row_idx, column=9, value=tolerance)
            
            # 是否通过（差值 <= 容差）
            if result_undist is not None:
                passed = diff <= tolerance
                status = "通过" if passed else "不通过"
                ws.cell(row=row_idx, column=10, value=status)
                # 设置背景色：通过=绿色，不通过=红色
                fill_color = "C6EFCE" if passed else "FFC7CE"
                ws.cell(row=row_idx, column=10).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            else:
                ws.cell(row=row_idx, column=10, value="--")
            
            # 设置对齐
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                if col not in [2, 3]:  # 图片列不需要设置对齐
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=len(self._all_results) + 1, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
        
        wb.save(file_path)


# ------------------------------ 主窗口 ------------------------------
class ZhangCalibrationApp(QMainWindow):
    """主窗口：Fluent 风格，顶部 Tab，内容区 7:3 / 3:1，全局间距与内边距。"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("张氏相机标定 - Zhang Calibration")
        self.setMinimumSize(1200, 740)
        self.setStyleSheet(f"QMainWindow {{ background-color: {BG_COLOR}; }}")
        icon_path = _app_icon_path()
        if icon_path:
            self.setWindowIcon(QIcon(icon_path))

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(CONTAINER_MARGIN, CONTAINER_MARGIN, CONTAINER_MARGIN, CONTAINER_MARGIN)
        main_layout.setSpacing(LAYOUT_SPACING)

        # 顶部 Tab 式功能选择（当前模式有高亮）
        tab_bar = QHBoxLayout()
        tab_bar.setSpacing(0)
        self._btn_group = QButtonGroup(self)
        self._btn_calib = PushButton("标定与结果") if _FLUENT_AVAILABLE else QPushButton("标定与结果")
        self._btn_calib.setCheckable(True)
        self._btn_calib.setChecked(True)
        self._btn_calib.setCursor(Qt.PointingHandCursor)
        self._btn_calib.setStyleSheet(
            "QPushButton { min-width: 140px; padding: 10px 20px; border: none; border-radius: 8px; "
            "background-color: rgb(50,50,50); color: #c0c0c0; } "
            "QPushButton:hover { background-color: rgb(60,60,60); color: #e0e0e0; } "
            "QPushButton:checked { background-color: #5e5cc9; color: white; font-weight: 600; }"
        )
        self._btn_validation = PushButton("畸变校正验证") if _FLUENT_AVAILABLE else QPushButton("畸变校正验证")
        self._btn_validation.setCheckable(True)
        self._btn_validation.setCursor(Qt.PointingHandCursor)
        self._btn_validation.setStyleSheet(
            "QPushButton { min-width: 140px; padding: 10px 20px; border: none; border-radius: 8px; "
            "background-color: rgb(50,50,50); color: #c0c0c0; } "
            "QPushButton:hover { background-color: rgb(60,60,60); color: #e0e0e0; } "
            "QPushButton:checked { background-color: #5e5cc9; color: white; font-weight: 600; }"
        )
        self._btn_group.addButton(self._btn_calib)
        self._btn_group.addButton(self._btn_validation)
        tab_bar.addWidget(self._btn_calib)
        tab_bar.addWidget(self._btn_validation)
        tab_bar.addStretch(1)
        main_layout.addLayout(tab_bar)

        self._stack = QStackedWidget()
        self._page_calib = CalibrationAndResultsPage(self)
        self._page_validation = UndistortionValidationPage(self)
        self._stack.addWidget(self._page_calib)
        self._stack.addWidget(self._page_validation)
        main_layout.addWidget(self._stack, stretch=1)

        self._btn_calib.clicked.connect(lambda: self._stack.setCurrentIndex(0))
        self._btn_validation.clicked.connect(lambda: self._stack.setCurrentIndex(1))


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    _apply_fluent_style(app)
    icon_path = _app_icon_path()
    if icon_path:
        app.setWindowIcon(QIcon(icon_path))
    win = ZhangCalibrationApp()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
